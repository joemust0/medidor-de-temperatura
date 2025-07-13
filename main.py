from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook, load_workbook
from tkinter import Tk, Button
import time
import os
from datetime import datetime

# Caminho do driver (ajuste conforme necessário)
CAMINHO_DRIVER = "chromedriver.exe"  # ou o caminho completo

# Função principal que coleta os dados
def coletar_dados_clima():
    
    servico = Service(CAMINHO_DRIVER)
    navegador = webdriver.Chrome(service=servico)

    
    navegador.get("https://www.climatempo.com.br/previsao-do-tempo/cidade/558/saopaulo-sp")
    time.sleep(2) 

    try:
        # Temperaturas
        temp_min = navegador.find_element(By.ID, "min-temp-1").text
        temp_max = navegador.find_element(By.ID, "max-temp-1").text

        # Umidades
        umid_min = navegador.find_element(
            By.XPATH, '//*[@id="mainContent"]/div[7]/div[3]/div[1]/div[2]/div[2]/div[3]/div[1]/ul/li[4]/div/p/span[1]'
        ).text.strip()

        umid_max = navegador.find_element(
            By.XPATH, '//*[@id="mainContent"]/div[7]/div[3]/div[1]/div[2]/div[2]/div[3]/div[1]/ul/li[4]/div/p/span[2]'
        ).text.strip()

        print(f"Temperatura: {temp_min} / {temp_max}")
        print(f"Umidade: {umid_min} / {umid_max}")

    except Exception as e:
        temp_min = temp_max = umid_min = umid_max = "Erro"
        print("Erro ao coletar dados:", e)

    navegador.quit()

    
    salvar_em_excel(temp_min, temp_max, umid_min, umid_max)

# Função que salva no Excel
def salvar_em_excel(temp_min, temp_max, umid_min, umid_max):
    nome_arquivo = "clima_sao_paulo.xlsx"
    cidade = "São Paulo"
    data = datetime.now().strftime("%d/%m/%Y %H:%M")

    if os.path.exists(nome_arquivo):
        wb = load_workbook(nome_arquivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Data/Hora", "Cidade", "Temp Mínima", "Temp Máxima", "Umid. Mínima", "Umid. Máxima"])

    ws.append([data, cidade, temp_min, temp_max, umid_min, umid_max])
    wb.save(nome_arquivo)
    print("Dados salvos com sucesso no Excel!")

# Interface gráfica com Tkinter
def criar_interface():
    janela = Tk()
    janela.title("Clima São Paulo")
    janela.geometry("300x150")

    botao = Button(janela, text="Gerar Clima", command=coletar_dados_clima)
    botao.pack(pady=50)

    janela.mainloop()

if __name__ == "__main__":
    criar_interface()
